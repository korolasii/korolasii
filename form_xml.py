import openpyxl
import xml.etree.ElementTree as ET

def format(file_path, output_xml, price_file = "Cost.xlsx"):
    gods_file = 'gods.xml'

    keywords = ['(Tabletki.ua)']

    tree = ET.parse(gods_file)
    root = tree.getroot()
    goods_data = {}
    for offer in root.findall('Offer'):
        code = offer.get('Code')
        name = offer.get('Name') or ''
        producer = offer.get('Producer') or ''
        if code:
            goods_data[code] = {
                'name': name,
                'producer': producer,
            }
            
    barcode_price_map = {}
    if price_file:
        price_workbook = openpyxl.load_workbook(price_file, data_only=True)
        for sheet in price_workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, values_only=True):
                try:
                    barcode = str(row[1]).strip()  
                    price = str(row[8]).strip()   
                    if barcode and price:
                        barcode_price_map[barcode] = price
                except Exception:
                    continue

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    results = []
    not_found_codes = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        header_row = None
        header_map = {}

        for row in sheet.iter_rows(min_row=1, max_col=10, values_only=True):
            headers = [str(cell).strip() if cell else "" for cell in row]
            if "Товар" in headers and "Штрих-код" in headers and "Ціна" in headers:
                header_row = headers
                break

        if not header_row:
            print(f"Заголовки не найдены на листе: {sheet_name}")
            continue

        for idx, col in enumerate(header_row):
            col = col.strip()
            if col == "Товар":
                header_map['name'] = idx
            elif col == "Штрих-код":
                header_map['barcode'] = idx
            elif col == "К-сть":
                header_map['quantity'] = idx
            elif col == "Ціна":
                header_map['price'] = idx

        reading = False
        for row in sheet.iter_rows(min_row=1, max_col=10, values_only=True):
            if row[header_map['name']] == "Товар":
                reading = True
                continue
            if not reading:
                continue

            try:
                name = str(row[header_map['name']]) if row[header_map['name']] else ''
                barcode = str(row[header_map['barcode']]) if row[header_map['barcode']] else ''
                quantity = str(row[header_map['quantity']]) if row[header_map['quantity']] else '0'
                price = barcode_price_map.get(barcode)
                if not price:
                    price = str(row[header_map['price']]) if row[header_map['price']] else '0'

                if any(word.lower() in name.lower() for word in keywords):
                    results.append([name, barcode, quantity, price])
            except Exception as e:
                continue

    root = ET.Element('Offers')
    for item in results:
        name, barcode, quantity, price = item
        if barcode in goods_data:
            data = goods_data[barcode]
            offer = ET.SubElement(root, 'Offer')
            offer.set('Code', barcode)
            offer.set('Name', data['name'])
            offer.set('Producer', data['producer'])
            offer.set('Tax', '0')
            offer.set('Price', price)
            offer.set('Quantity', quantity)
            offer.set('PriceReserve', price)
            offer.set('Barcode', barcode)
        else:
            not_found_codes.append(barcode)

    tree = ET.ElementTree(root)
    tree.write(output_xml, encoding='utf-8', xml_declaration=True)
    print(f'✅ XML-файл создан: {output_xml}')

    if not_found_codes:
        print('⚠️ Штрихкоды не найдены в gods.xml:')
        for code in set(not_found_codes):
            print(code)
    else:
        print('✅ Все штрихкоды найдены в gods.xml.')

def main_format(file_path, output_xml):
    format(file_path, output_xml)

if __name__ == "__main__":
    format("downloads/Люстдорфська.xlsx", 'Люстдорфська.xml')
    format("downloads/Троїцька (2).xlsx", 'Троїцька.xml')
